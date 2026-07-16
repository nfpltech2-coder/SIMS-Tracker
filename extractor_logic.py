"""
SIMS PDF Extraction Logic — Tab 2
Extracts SIM Number, Quantity, HS Code from SIMS Application PDFs.
Generates a 'Merged to Compare' key for downstream matching.
"""
import os
import re
import decimal
from typing import Dict, List
from datetime import datetime

import pdfplumber
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ── Excel Styling ──
FILL_HEADER = PatternFill(start_color="0056B3", end_color="0056B3", fill_type="solid")
FONT_HEADER = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def format_qty_for_merge(qty) -> str:
    """Rounds to 3 decimal places (ROUND_HALF_UP), strips trailing zeros."""
    try:
        if pd.isna(qty) or str(qty).strip() == "":
            return ""
        d = decimal.Decimal(str(qty))
        rounded = d.quantize(decimal.Decimal('0.001'), rounding=decimal.ROUND_HALF_UP)
        q_str = f"{rounded:f}"
        if '.' in q_str:
            q_str = q_str.rstrip('0').rstrip('.')
        return q_str
    except (ValueError, TypeError, decimal.InvalidOperation):
        return str(qty).strip()


def extract_pdf_data(path: str) -> Dict[str, str]:
    """Extracts SIM Number, Quantity, HS Code from a single SIMS PDF."""
    try:
        with pdfplumber.open(path) as pdf:
            text = ""
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"

        sim_match = re.search(r"SIMS Number\s+([\w]+)", text, re.IGNORECASE)
        date_match = re.search(r"SIMS Date\s+(\d{2}/\d{2}/\d{4})", text, re.IGNORECASE)
        qty_match = re.search(r"Quantity\s+([\d\.]+)", text, re.IGNORECASE)
        hs_match = re.search(r"HS Code\s+([\d]+)", text, re.IGNORECASE)
        coo_match = re.search(
            r"Country of Origin\s+([A-Z\s]+?)(?=\r?\n|Country of Consignment|$)",
            text, re.IGNORECASE
        )

        sim_number = sim_match.group(1).strip() if sim_match else "Not Found"
        sim_date = None
        if date_match:
            try:
                sim_date = datetime.strptime(date_match.group(1).strip(), "%d/%m/%Y")
            except:
                pass
        
        quantity = qty_match.group(1).strip() if qty_match else "Not Found"
        hs_code = hs_match.group(1).strip() if hs_match else "Not Found"
        country_origin = coo_match.group(1).strip() if coo_match else "Not Found"

        merged = ""
        if hs_code != "Not Found" and quantity != "Not Found":
            merged = f"{hs_code}/{format_qty_for_merge(quantity)}"

        return {
            "File Name": os.path.basename(path),
            "SIM Number": sim_number,
            "SIMS Date": sim_date,
            "Quantity": quantity,
            "HS Code": hs_code,
            "Country of Origin": country_origin,
            "Merged to Compare": merged,
        }
    except Exception as e:
        return {
            "File Name": os.path.basename(path),
            "SIM Number": f"Error: {e}",
            "SIMS Date": "Error",
            "Quantity": "Error",
            "HS Code": "Error",
            "Country of Origin": "Error",
            "Merged to Compare": "Error",
        }


def polish_excel_sheet(ws):
    """Apply Nagarkot header styling to an openpyxl worksheet."""
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    for col in ws.columns:
        header_val = str(col[0].value)
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
        
        # If it's the SIMS Date column, apply date format
        if header_val == "SIMS Date":
            for cell in col[1:]: # Skip header
                cell.number_format = 'dd-mmm-yyyy'
                cell.alignment = ALIGN_CENTER


def extract_payment_slip_data(path: str) -> Dict[str, any]:
    """Extracts Amount and Transaction Date from payment receipt PDFs."""
    try:
        with pdfplumber.open(path) as pdf:
            text = ""
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"
        
        # 1. Extract Amount
        amount = 0.0
        # Layout A: E-Acknowledgement format (e.g. "Amount 750.00")
        amt_match = re.search(r"Amount\s+([\d\.]+)", text, re.IGNORECASE)
        if amt_match:
            amount = float(amt_match.group(1).strip())
        else:
            # Layout B: Bharatkosh format (e.g. "sum of INR 3000" or just "INR 3000")
            bk_amt_match = re.search(r"sum of INR\s*([\d\.,]+)", text, re.IGNORECASE)
            if not bk_amt_match:
                bk_amt_match = re.search(r"INR\s*([\d\.,]+)", text, re.IGNORECASE)
            if bk_amt_match:
                amount_str = bk_amt_match.group(1).replace(",", "").strip()
                amount = float(amount_str)
        
        # 2. Extract Date
        date_str = "Not Found"
        # Layout A: E-Acknowledgement format (e.g. "Transaction Date 12-05-2026")
        date_match = re.search(r"Transaction Date\s+(\d{2}-\d{2}-\d{4})", text, re.IGNORECASE)
        if date_match:
            try:
                raw_date = date_match.group(1).strip()
                dt = datetime.strptime(raw_date, "%d-%m-%Y")
                date_str = dt.strftime("%d-%b-%Y").upper() # e.g. 12-MAY-2026
            except:
                date_str = date_match.group(1).strip()
        else:
            # Layout B: Bharatkosh format (e.g. "Dated: Jul 7 2026" or "Dated Jul 7 2026")
            bk_date_match = re.search(r"Dated:?\s*([A-Za-z]{3}\s+\d{1,2}\s+\d{4})", text, re.IGNORECASE)
            if bk_date_match:
                try:
                    raw_date = re.sub(r'\s+', ' ', bk_date_match.group(1)).strip()
                    dt = datetime.strptime(raw_date, "%b %d %Y")
                    date_str = dt.strftime("%d-%b-%Y").upper()
                except:
                    date_str = bk_date_match.group(1).strip()
        
        return {
            "Amount": amount,
            "Date": date_str,
            "File": os.path.basename(path)
        }
    except Exception as e:
        return {"Amount": 0.0, "Date": f"Error: {e}", "File": os.path.basename(path)}


def run_extraction(
    pdf_paths: List[str],
    output_path: str,
    progress_callback=None,
    done_callback=None,
    error_callback=None,
):
    """Extracts data from a list of PDFs and writes to an Excel file."""
    try:
        results = []
        total = len(pdf_paths)
        for i, fpath in enumerate(pdf_paths):
            if progress_callback:
                pct = ((i) / total) * 100
                progress_callback(pct, f"Extracting: {os.path.basename(fpath)}")
            results.append(extract_pdf_data(fpath))

        df = pd.DataFrame(results)
        df = df[["SIM Number", "SIMS Date", "Quantity", "HS Code", "Country of Origin", "Merged to Compare", "File Name"]]

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="SIMS Data")
            polish_excel_sheet(writer.sheets["SIMS Data"])

        if progress_callback:
            progress_callback(100, "Extraction complete!")
        if done_callback:
            done_callback(output_path)

    except Exception as e:
        if error_callback:
            error_callback(str(e))
