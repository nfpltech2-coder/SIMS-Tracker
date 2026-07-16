import pandas as pd
import os
from datetime import datetime

def clean_excel_formula(val):
    """Removes Excel formula wrappers like '="VALUE"'."""
    s = str(val).strip()
    if s.startswith('="') and s.endswith('"'):
        return s[2:-1]
    return val

def read_any_df(path):
    """Helper to read Excel or CSV with auto-delimiter detection and formula cleanup."""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xlsx':
        df = pd.read_excel(path, keep_default_na=False)
    else:
        # Try common delimiters for CSV
        df = None
        for sep in [',', ';', '\t']:
            try:
                temp_df = pd.read_csv(path, sep=sep, keep_default_na=False, low_memory=False)
                if len(temp_df.columns) > 1:
                    df = temp_df
                    break
            except:
                continue
        if df is None:
            df = pd.read_csv(path, keep_default_na=False, low_memory=False)
    
    # Apply cleanup to all object columns
    for col in df.select_dtypes(include=['object']):
        df[col] = df[col].apply(clean_excel_formula)
    return df

def process_tracker_files(job_no, master_path, invoice_path, output_dir, update_callback=None, error_callback=None, shakti_push_callback=None, done_callback=None):
    try:
        if update_callback: update_callback("Loading files...")
        master_df = read_any_df(master_path)
        invoice_df = read_any_df(invoice_path)
        
        # --- TRUNCATE AT SUMMARY ROWS ---
        # Look for "Total", "Unique", or completely empty rows to stop processing
        stop_idx = None
        for i in range(len(invoice_df)):
            row_vals = [str(v).lower().strip() for v in invoice_df.iloc[i].values]
            # Check first column for keywords
            if any(kw in row_vals[0] for kw in ["total", "unique"]):
                stop_idx = i
                break
            # Check if entire row is empty
            if all(v in ["", "nan", "none"] for v in row_vals):
                stop_idx = i
                break
        
        if stop_idx is not None:
            invoice_df = invoice_df.iloc[:stop_idx].copy()
        
        if update_callback: update_callback("Mapping CTH from Master...")
        
        master_part_col = 'PartNo' if 'PartNo' in master_df.columns else None
        if not master_part_col:
            # Check for other master part col names if needed
            for col in ['PartNo', 'Part Number', 'Material']:
                if col in master_df.columns:
                    master_part_col = col
                    break
        
        if not master_part_col:
            raise ValueError("Master sheet is missing 'PartNo' column.")
            
        inv_part_col = None
        # Support various names for Part Number
        part_cols = ['Part No.', 'Mat. NO.', 'Part Number', 'PART NO', 'Customer P/N', 'Foton P/N', 'Article No']
        for col in part_cols:
            if col in invoice_df.columns:
                inv_part_col = col
                break
        
        if not inv_part_col:
            raise ValueError(f"Invoice sheet missing Part Number column. (Checked: {', '.join(part_cols)})")
            
        # Clean columns for matching
        master_df[master_part_col] = master_df[master_part_col].astype(str).str.strip()
        cth_map = master_df.dropna(subset=['CTH1']).set_index(master_part_col)['CTH1'].to_dict()
        # Create a version of the map with leading zeros removed from keys for fuzzy matching
        cth_map_stripped = {str(k).lstrip('0'): v for k, v in cth_map.items()}
        
        invoice_df[inv_part_col] = invoice_df[inv_part_col].astype(str).str.strip()
        
        def smart_match_cth(part):
            p = str(part).strip()
            if not p: return None, 'N/A'
            # 1. Try exact match
            if p in cth_map: return cth_map[p], 'Master'
            # 2. Try match without leading zeros (handles 098767 -> 98767)
            p_stripped = p.lstrip('0')
            if p_stripped in cth_map_stripped:
                return cth_map_stripped[p_stripped], 'Master (Matched but zero was removed)'
            return None, 'Invoice (Fallback)'

        matches = invoice_df[inv_part_col].apply(smart_match_cth)
        invoice_df['AS PER DHL'] = matches.apply(lambda x: x[0]).astype(object)
        invoice_df['CTH_Source'] = matches.apply(lambda x: x[1])
        
        # --- Fallback logic for HS Code ---
        # Find HS Code col in invoice
        inv_hs_col = None
        for col in ['HS Code', 'HS-CODE', 'HSCODE']:
            if col in invoice_df.columns:
                inv_hs_col = col
                break
        
        if inv_hs_col:
            # For rows where master failed, take from invoice
            mask = invoice_df['AS PER DHL'].isna() | (invoice_df['AS PER DHL'].astype(str).str.lower() == 'nan') | (invoice_df['AS PER DHL'] == '')
            invoice_df.loc[mask, 'AS PER DHL'] = invoice_df.loc[mask, inv_hs_col]
            invoice_df.loc[mask, 'CTH_Source'] = 'Invoice (Fallback)'

        def get_sqc_with_calc(row):
            def clean_num(v):
                if pd.isna(v) or v == '': return 0.0
                try: return float(str(v).replace(',', '').replace(' ', '').replace('kg', '').strip())
                except: return 0.0

            if 'Weight per PC (kg)' in row and pd.notna(row['Weight per PC (kg)']) and str(row['Weight per PC (kg)']).strip() != '':
                qty_col = 'Quantity' if 'Quantity' in row else 'Qty' if 'Qty' in row else None
                w_pc = clean_num(row['Weight per PC (kg)'])
                qty = clean_num(row.get(qty_col, 1))
                return w_pc * qty, f"{w_pc} * {qty}"
            
            for col in ['Net Weight', 'NET WEIGHT(KG)', 'Net wt', 'Quantity (EA)']:
                if col in row and pd.notna(row[col]) and str(row[col]).strip() != '':
                    val = clean_num(row[col])
                    return val, f"Direct: {col}"
            
            return 0.0, ""

        sqc_results = invoice_df.apply(get_sqc_with_calc, axis=1)
        invoice_df['SQC_Qty'] = sqc_results.apply(lambda x: x[0])
        invoice_df['SQC Calculation'] = sqc_results.apply(lambda x: x[1])
        invoice_df['Bot Calculation Check'] = invoice_df['SQC_Qty'] / 1000

        if update_callback: update_callback("Routing records...")
        
        df_routed = invoice_df.copy()
        
        # Function to safely get prefixes
        def get_prefix(val, length):
            val_str = str(val).split('.')[0].strip() # handle floats like 87089900.0
            if val_str.lower() in ['nan', 'none', 'null', '']:
                return ''
            return val_str[:length]
            
        df_routed['CTH_PREFIX_2'] = df_routed['AS PER DHL'].apply(lambda x: get_prefix(x, 2))
        df_routed['CTH_PREFIX_4'] = df_routed['AS PER DHL'].apply(lambda x: get_prefix(x, 4))
        
        sheet_dict = {
            'Sheet1': [],
            'PIMS': [],
            'Copper': [],      # Prefix 74
            'Aluminium': [],   # Prefix 76
            'Perfumery': [],    # Prefix 33
            'Tyre': [],
            'Radar Radiance': [],
            'Camera': []       # Prefix 8525
        }
        
        csv_items = []
        
        for idx, row in df_routed.iterrows():
            prefix_2 = row['CTH_PREFIX_2']
            prefix_4 = row['CTH_PREFIX_4']
            
            if prefix_2 == '48':
                sheet_dict['PIMS'].append(row)
            elif prefix_2 == '74':
                sheet_dict['Copper'].append(row)
            elif prefix_2 == '76':
                sheet_dict['Aluminium'].append(row)
            elif prefix_2 == '33':
                sheet_dict['Perfumery'].append(row)
            elif prefix_4 == '4011':
                sheet_dict['Tyre'].append(row)
            elif prefix_4 == '8526':
                sheet_dict['Radar Radiance'].append(row)
            elif prefix_4 == '8525':
                sheet_dict['Camera'].append(row)
            elif prefix_2 in ['72', '73', '86']:
                sheet_dict['Sheet1'].append(row)
                csv_items.append(row)
            else:
                sheet_dict['Sheet1'].append(row)
                
        csv_payload_items = []
        df_csv = None
        
        if csv_items:
            df_csv_raw = pd.DataFrame(csv_items)
            
            df_csv = pd.DataFrame()
            inv_col = None
            for col in ['Invoice Number', 'Invoice No', 'Inv no']:
                if col in df_csv_raw.columns:
                    inv_col = col
                    break
            if not inv_col: inv_col = df_csv_raw.columns[0]

            desc_col = None
            for col in ['Description', 'DESC', 'Item Description']:
                if col in df_csv_raw.columns:
                    desc_col = col
                    break

            df_csv['Inv No'] = df_csv_raw[inv_col]
            df_csv['CTH'] = df_csv_raw['AS PER DHL'].astype(str).str.split('.').str[0]
            df_csv['Model'] = df_csv_raw[inv_part_col]
            df_csv['SQC_Qty'] = df_csv_raw['SQC_Qty']
            
            desc_series = df_csv_raw[desc_col] if desc_col else pd.Series([""]*len(df_csv_raw), index=df_csv_raw.index)
            
            for idx, r in df_csv.iterrows():
                csv_payload_items.append({
                    "Inv_No": str(r.get('Inv No', '')),
                    "CTH": str(r.get('CTH', '')),
                    "Model": str(r.get('Model', '')),
                    "Description": str(desc_series.loc[idx]),
                    "SQC_Qty": str(r.get('SQC_Qty', '')),
                    "SIMS_Code": ""
                })
        
        # --- SHAKTI PUSH BEFORE SAVING ---
        shakti_msg = "Skipped"
        if shakti_push_callback:
            if update_callback: update_callback("Pushing to Shakti...")
            ok, msg = shakti_push_callback(csv_payload_items)
            if not ok:
                if error_callback: error_callback(f"Shakti Push Failed:\n{msg}\n\nFiles were NOT generated.")
                return
            shakti_msg = msg

        # --- NOW SAVE FILES ---
        if update_callback: update_callback("Generating outputs...")
        excel_out_path = os.path.join(output_dir, f"{job_no}_reference.xlsx")
        
        from openpyxl.styles import PatternFill
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        with pd.ExcelWriter(excel_out_path, engine='openpyxl') as writer:
            for sheet_name, rows in sheet_dict.items():
                if rows:
                    df_sheet = pd.DataFrame(rows).reset_index(drop=True)
                    fallback_indices = df_sheet[df_sheet['CTH_Source'] == 'Invoice (Fallback)'].index
                    df_sheet = df_sheet.drop(columns=['CTH_PREFIX_2', 'CTH_PREFIX_4'], errors='ignore')
                    df_sheet.to_excel(writer, index=False, sheet_name=sheet_name)
                    
                    worksheet = writer.sheets[sheet_name]
                    as_per_dhl_idx = None
                    for col_idx, col in enumerate(df_sheet.columns):
                        if col == 'AS PER DHL':
                            as_per_dhl_idx = col_idx + 1
                            break
                    
                    if as_per_dhl_idx:
                        # Use sequential index from the sheet, not original dataframe index
                        for row_pos in fallback_indices:
                            worksheet.cell(row=row_pos + 2, column=as_per_dhl_idx).fill = yellow_fill
                            
        if not any(sheet_dict.values()):
            with pd.ExcelWriter(excel_out_path, engine='openpyxl') as writer:
               pd.DataFrame(columns=invoice_df.columns).to_excel(writer, index=False, sheet_name='Sheet1')

        csv_out_path = None
        if csv_items and df_csv is not None:
            csv_out_path = os.path.join(output_dir, f"IR{job_no}.csv")
            df_csv.to_csv(csv_out_path, index=False)
            
        if done_callback: done_callback(excel_out_path, csv_out_path, shakti_msg)
            
    except Exception as e:
        if error_callback: error_callback(str(e))
